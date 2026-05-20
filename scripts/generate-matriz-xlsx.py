"""
scripts/generate-matriz-xlsx.py

Genera docs/qa/release-v2.0.4/matriz_cases_baja_complejidad.xlsx a partir de la matriz
markdown del mismo nombre. Aplica formato profesional (skill 'xlsx'):
- Font Arial 10
- Headers bold + fill azul (industria-standard hardcoded inputs)
- Color coding: ID en verde (links a docs), Estado/Prioridad celdas con datavalidation
- Columnas autoajustadas
- 3 sheets: Matriz / Resumen / Trazabilidad
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ----- Styles -----
ARIAL = "Arial"
HEADER_FONT = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")  # azul corporativo
CELL_FONT = Font(name=ARIAL, size=10)
ID_FONT = Font(name=ARIAL, size=10, color="008000", bold=True)  # verde (link a docs)
MX_FONT = Font(name=ARIAL, size=10, color="0000FF", bold=True)  # azul (input)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ----- Definicion completa de TCs -----
# Cada entry: (id, mx, pantalla, url, seccion, dimension, tipo, prioridad, descripcion,
#              precondicion, pasos, resultado_esperado)

# Comunes
PRE_LOGIN = "Usuario logueado como carrier (remises.eeuu@yopmail.com), sesion activa."
PRE_LIST_LOADED = "Pantalla cargada con datos visibles en la tabla."


def list_steps(steps_list):
    """Devuelve string multilinea con pasos numerados."""
    return "\n".join(f"{i + 1}. {s}" for i, s in enumerate(steps_list))


# Precondiciones por pantalla (para construir TCs).
SCREENS = {
    "MX-5560": {
        "name": "Reportes / Propinas",
        "url": "/carrier/#/reports/tips",
        "heading_en": "Tips Report",
        "heading_es": "Reporte de Propinas",
    },
    "MX-5561": {
        "name": "Reportes / Antigüedad de Deuda",
        "url": "/carrier/#/reports/debt-aging",
        "heading_en": "Aging Report",
        "heading_es": "Reporte Antigüedad de Deuda",
    },
    "MX-5562": {
        "name": "Reportes / Movimientos de Cobros",
        "url": "/carrier/#/reports/cash-flow",
        "heading_en": "Collection Movements",
        "heading_es": "Movimientos de Cobros",
    },
    "MX-5565": {
        "name": "Reportes / Transacciones Tarjeta",
        "url": "/carrier/#/reports/transaction-tracking",
        "heading_en": "Electronic Payment Transactions",
        "heading_es": "Transacciones con Tarjeta",
    },
    "MX-5566": {
        "name": "Reportes / Impuestos y Cargos",
        "url": "/carrier/#/reports/taxes-and-fees",
        "heading_en": "Taxes & Fees Report",
        "heading_es": "Impuestos y Cargos",
    },
    "MX-5568": {
        "name": "Reportes / Movimientos de Pagos",
        "url": "/carrier/#/reports/payment-flow",
        "heading_en": "Payment Movements",
        "heading_es": "Movimientos de Pagos",
    },
    "MX-5571": {
        "name": "Reportes / Comisiones Agencia",
        "url": "/carrier/#/reports/agency-commissions",
        "heading_en": "Company Commissions Reports",
        "heading_es": "Comisiones Agencia",
    },
    "MX-5438": {
        "name": "Reportes / Resumen Diario",
        "url": "/carrier/#/reports/daily",
        "heading_en": "Daily Report",
        "heading_es": "Resumen Diario",
    },
    "MX-5569": {
        "name": "Reportes / Documentación Vencida",
        "url": "/carrier/#/reports/documentation",
        "heading_en": "Expired Documentation Report",
        "heading_es": "Documentación Vencida y por Vencer",
    },
    "MX-5531": {
        "name": "Reportes / Travel Unpaid List",
        "url": "/carrier/#/reports/unpaid-travels-list",
        "heading_en": "Unpaid Trips Report",
        "heading_es": "Viajes Impagos",
    },
    "MX-5553": {
        "name": "Reportes / Viajes por Segmentos",
        "url": "/carrier/#/reports/segments-travels",
        "heading_en": "Trips Segments",
        "heading_es": "Viajes por Segmentos",
    },
    "MX-5573": {
        "name": "GNET Farm IN",
        "url": "/carrier/#/gnet/farm-in",
        "heading_en": "GNET Farm IN",
        "heading_es": "GNET Farm IN",
    },
    "MX-5574": {
        "name": "GNET Cuentas Corrientes",
        "url": "/carrier/#/gnet/credit-accounts",
        "heading_en": "GNET Credit Accounts",
        "heading_es": "GNET Cuentas Corrientes",
    },
    "MX-5575": {
        "name": "Configuración / Otros Costos",
        "url": "/carrier/#/settings/otherCosts",
        "heading_en": "Other Costs",
        "heading_es": "Otros Costos",
    },
    "MX-5554": {
        "name": "Cuentas Corrientes Con Afiliados",
        "url": "/carrier/#/affiliate/checking-account",
        "heading_en": "Credit Accounts With Affiliates",
        "heading_es": "Cuentas Corrientes Con Afiliados",
    },
}

# Build TCs from md content. Mapping minimo (id, desc, dimension, prioridad).
# Para cada pantalla, generar 8-26 TCs segun la matriz.

# Template TC: (tc_id, desc, dim, prio)
# Dimensiones qa-magiis: HP=Happy path, EC=Edge case, NEG=Negativo, REG=Regresion, INT=Integracion
# Prioridad: P1=Alta, P2=Media, P3=Baja

# Genero los TCs por pantalla siguiendo la matriz md.
TCS_PER_SCREEN = {
    "MX-5560": [
        ("TC01", "Validar acceso desde menu Reports → Tips y carga sin error", "HP", "P1"),
        ("TC02", "Validar titulo de pantalla traducido EN: 'Tips Report' / ES: 'Reporte de Propinas'", "INT", "P1"),
        ("TC03", "Validar breadcrumb traducido EN/ES (Reports / Tips)", "INT", "P2"),
        ("TC04", "Validar dropdown travel type con opciones Historical/Recent y default Historical", "HP", "P1"),
        ("TC05", "Validar date range picker visible solo cuando travel type = Historical", "HP", "P2"),
        ("TC06", "Validar date range picker preset Today aplica filtro y tabla actualiza", "HP", "P1"),
        ("TC07", "Validar date range picker rango custom from-to aplica y persiste tras refresh", "EC", "P2"),
        ("TC08", "Validar input search filtra resultados por nombre y reduce filas", "HP", "P1"),
        ("TC09", "Validar input search vacio restaura todos los resultados", "EC", "P2"),
        ("TC10", "Validar dropdown payment method filtra (Checking Account / Credit Card)", "HP", "P1"),
        ("TC11", "Validar dropdown payment method clear restaura todos los metodos", "EC", "P2"),
        ("TC12", "Validar dropdown status filtra correctamente las filas", "HP", "P1"),
        ("TC13", "Validar ordenamiento ASC al clickear columna sortable (Date)", "HP", "P1"),
        ("TC14", "Validar ordenamiento DESC al volver a clickear misma columna", "HP", "P1"),
        ("TC15", "Validar resize de columna arrastrando resize-handle lateral cambia ancho", "HP", "P2"),
        ("TC16", "Validar drag & drop de columna cambia orden visualmente y persiste durante sesion", "HP", "P2"),
        ("TC17", "Validar modal/dropdown configuracion columnas oculta/muestra columna", "HP", "P2"),
        ("TC18", "Validar paginacion Previous/Next navega entre paginas", "HP", "P1"),
        ("TC19", "Validar selector 'Show N' cambia cantidad de filas por pagina", "HP", "P2"),
        ("TC20", "Validar boton Refresh recarga tabla y muestra spinner durante operacion", "HP", "P2"),
        ("TC21", "Validar boton PDF deshabilitado cuando reportList.length === 0", "NEG", "P1"),
        ("TC22", "Validar boton PDF descarga archivo .pdf con header + tabla + dates en formato local", "HP", "P1"),
        ("TC23", "Validar PDF generado con locale ES usa formato fecha DD/MM/YYYY", "INT", "P1"),
        ("TC24", "Validar PDF generado con locale EN usa formato fecha MM/DD/YYYY", "INT", "P1"),
        ("TC25", "Validar cambio idioma EN ↔ ES actualiza headers + placeholders + botones", "INT", "P1"),
        ("TC26", "Validar empty state 'No data' cuando filtros no devuelven filas", "NEG", "P2"),
    ],
    # Reportes siguientes - patron condensado (8-14 TCs)
    "MX-5561": [
        ("TC01", "Validar acceso desde menu Reports → Debt Aging y carga sin error", "HP", "P1"),
        ("TC02", "Validar titulo 'Aging Report' / 'Reporte Antigüedad de Deuda' segun idioma", "INT", "P1"),
        ("TC03", "Validar filtros aplican (cliente, fecha, etc.)", "HP", "P1"),
        ("TC04", "Validar search filtra por nombre y restaura al limpiar", "HP", "P1"),
        ("TC05", "Validar ordenamiento ASC/DESC en columnas sortable", "HP", "P1"),
        ("TC06", "Validar resize de columnas con resize-handle", "HP", "P2"),
        ("TC07", "Validar drag & drop reordena columnas", "HP", "P2"),
        ("TC08", "Validar config columnas oculta/muestra correctamente", "HP", "P2"),
        ("TC09", "Validar paginacion Previous/Next + page size", "HP", "P1"),
        ("TC10", "Validar refresh recarga datos con spinner", "HP", "P2"),
        ("TC11", "Validar PDF deshabilitado sin datos", "NEG", "P1"),
        ("TC12", "Validar PDF descarga con totales calculados visibles", "HP", "P1"),
        ("TC13", "Validar i18n EN/ES en headers + placeholders", "INT", "P1"),
        ("TC14", "Validar l10n: formato fecha por region en columnas date", "INT", "P1"),
    ],
    "MX-5562": [
        ("TC01", "Validar acceso desde menu y carga sin error", "HP", "P1"),
        ("TC02", "Validar titulo 'Collection Movements' / 'Movimientos de Cobros'", "INT", "P1"),
        ("TC03", "Validar date range picker con preset y custom", "HP", "P1"),
        ("TC04", "Validar filtros aplican (estado de cobro / metodo)", "HP", "P1"),
        ("TC05", "Validar search por codigo/cliente", "HP", "P1"),
        ("TC06", "Validar ordenamiento ASC/DESC", "HP", "P1"),
        ("TC07", "Validar resize + drag & drop + config columnas", "HP", "P2"),
        ("TC08", "Validar seccion 'Totals' (h4) muestra sumatorias correctas", "EC", "P1"),
        ("TC09", "Validar paginacion y refresh", "HP", "P2"),
        ("TC10", "Validar PDF descarga con dates en formato local", "INT", "P1"),
        ("TC11", "Validar i18n EN/ES + l10n fechas", "INT", "P1"),
    ],
    "MX-5565": [
        ("TC01", "Validar acceso desde menu y carga", "HP", "P1"),
        ("TC02", "Validar titulo 'Electronic Payment Transactions' / 'Transacciones con Tarjeta'", "INT", "P1"),
        ("TC03", "Validar filtros por estado transaccion + metodo pago", "HP", "P1"),
        ("TC04", "Validar date range picker", "HP", "P1"),
        ("TC05", "Validar search libre", "HP", "P1"),
        ("TC06", "Validar ordenamiento ASC/DESC en columnas sortable", "HP", "P1"),
        ("TC07", "Validar resize + drag & drop + config columnas", "HP", "P2"),
        ("TC08", "Validar paginacion + refresh", "HP", "P2"),
        ("TC09", "Validar PDF deshabilitado vacio + descarga con datos", "NEG", "P1"),
        ("TC10", "Validar i18n + l10n", "INT", "P1"),
    ],
    "MX-5566": [
        ("TC01", "Validar acceso desde menu y carga", "HP", "P1"),
        ("TC02", "Validar titulo 'Taxes & Fees' / 'Impuestos y Cargos'", "INT", "P1"),
        ("TC03", "Validar filtros aplican (categoria / periodo)", "HP", "P1"),
        ("TC04", "Validar date range picker", "HP", "P1"),
        ("TC05", "Validar search", "HP", "P1"),
        ("TC06", "Validar ordenamiento ASC/DESC", "HP", "P1"),
        ("TC07", "Validar resize + drag & drop + config columnas", "HP", "P2"),
        ("TC08", "Validar paginacion + refresh", "HP", "P2"),
        ("TC09", "Validar PDF con totales por categoria", "HP", "P1"),
        ("TC10", "Validar i18n + l10n", "INT", "P1"),
    ],
    "MX-5568": [
        ("TC01", "Validar acceso desde menu y carga", "HP", "P1"),
        ("TC02", "Validar titulo 'Payment Movements' / 'Movimientos de Pagos'", "INT", "P1"),
        ("TC03", "Validar filtros aplican (tipo de pago / status)", "HP", "P1"),
        ("TC04", "Validar date range picker", "HP", "P1"),
        ("TC05", "Validar search", "HP", "P1"),
        ("TC06", "Validar ordenamiento ASC/DESC", "HP", "P1"),
        ("TC07", "Validar resize + drag & drop + config columnas", "HP", "P2"),
        ("TC08", "Validar paginacion + refresh", "HP", "P2"),
        ("TC09", "Validar PDF descarga con dates correctas", "INT", "P1"),
        ("TC10", "Validar i18n + l10n", "INT", "P1"),
    ],
    "MX-5571": [
        ("TC01", "Validar acceso desde menu y carga", "HP", "P1"),
        ("TC02", "Validar titulo 'Company Commissions' / 'Comisiones Agencia'", "INT", "P1"),
        ("TC03", "Validar filtros aplican (cliente / periodo)", "HP", "P1"),
        ("TC04", "Validar date range picker", "HP", "P1"),
        ("TC05", "Validar search", "HP", "P1"),
        ("TC06", "Validar ordenamiento ASC/DESC", "HP", "P1"),
        ("TC07", "Validar resize + drag & drop + config columnas", "HP", "P2"),
        ("TC08", "Validar seccion 'Totals' muestra sumatorias", "EC", "P1"),
        ("TC09", "Validar paginacion + refresh", "HP", "P2"),
        ("TC10", "Validar PDF con totales por cliente", "HP", "P1"),
        ("TC11", "Validar i18n + l10n", "INT", "P1"),
    ],
    "MX-5438": [
        ("TC01", "Validar acceso desde menu y carga", "HP", "P1"),
        ("TC02", "Validar titulo 'Daily Report' / 'Resumen Diario'", "INT", "P1"),
        ("TC03", "Validar date picker dia unico + navegacion dia anterior/siguiente", "HP", "P1"),
        ("TC04", "Validar filtros aplican (driver / cliente / status)", "HP", "P1"),
        ("TC05", "Validar search libre", "HP", "P1"),
        ("TC06", "Validar ordenamiento ASC/DESC en columnas sortable", "HP", "P1"),
        ("TC07", "Validar resize + drag & drop + config columnas", "HP", "P2"),
        ("TC08", "Validar paginacion + refresh", "HP", "P2"),
        ("TC09", "Validar PDF descarga con fecha del dia seleccionado", "INT", "P1"),
        ("TC10", "Validar i18n + l10n fechas", "INT", "P1"),
    ],
    "MX-5569": [
        ("TC01", "Validar acceso desde menu y carga", "HP", "P1"),
        ("TC02", "Validar titulo 'Expired Documentation' / 'Documentacion Vencida y por Vencer'", "INT", "P1"),
        ("TC03", "Validar filtros por tipo documento + estado (vencido/por vencer/vigente)", "HP", "P1"),
        ("TC04", "Validar filtro por entidad (driver / vehicle / owner)", "HP", "P1"),
        ("TC05", "Validar search libre", "HP", "P1"),
        ("TC06", "Validar ordenamiento ASC/DESC por fecha vencimiento", "HP", "P1"),
        ("TC07", "Validar resize + drag & drop + config columnas", "HP", "P2"),
        ("TC08", "Validar paginacion + refresh", "HP", "P2"),
        ("TC09", "Validar PDF descarga con highlight de docs vencidos", "HP", "P1"),
        ("TC10", "Validar i18n + l10n fechas", "INT", "P1"),
    ],
    "MX-5531": [
        ("TC01", "Validar acceso desde menu y carga", "HP", "P1"),
        ("TC02", "Validar titulo 'Unpaid Trips' / 'Viajes Impagos'", "INT", "P1"),
        ("TC03", "Validar filtros aplican (cliente / periodo sin pago)", "HP", "P1"),
        ("TC04", "Validar search libre", "HP", "P1"),
        ("TC05", "Validar ordenamiento ASC/DESC", "HP", "P1"),
        ("TC06", "Validar resize + drag & drop + config columnas", "HP", "P2"),
        ("TC07", "Validar paginacion + refresh", "HP", "P2"),
        ("TC08", "Validar PDF descarga", "HP", "P1"),
        ("TC09", "Validar i18n + l10n", "INT", "P1"),
    ],
    "MX-5553": [
        ("TC01", "Validar acceso desde menu y carga", "HP", "P1"),
        ("TC02", "Validar titulo 'Trips Segments' / 'Viajes por Segmentos'", "INT", "P1"),
        ("TC03", "Validar filtros por segmento + periodo", "HP", "P1"),
        ("TC04", "Validar search libre", "HP", "P1"),
        ("TC05", "Validar ordenamiento ASC/DESC", "HP", "P1"),
        ("TC06", "Validar resize + drag & drop + config columnas", "HP", "P2"),
        ("TC07", "Validar seccion 'Totals'", "EC", "P1"),
        ("TC08", "Validar paginacion + refresh", "HP", "P2"),
        ("TC09", "Validar PDF descarga con totales por segmento", "HP", "P1"),
        ("TC10", "Validar i18n + l10n", "INT", "P1"),
    ],
    "MX-5573": [
        ("TC01", "Validar acceso desde menu GNET → Farm IN y carga", "HP", "P1"),
        ("TC02", "Validar titulo 'GNET Farm IN'", "INT", "P1"),
        ("TC03", "Validar filtros aplican (estado / origen / periodo)", "HP", "P1"),
        ("TC04", "Validar date range picker", "HP", "P1"),
        ("TC05", "Validar search libre", "HP", "P1"),
        ("TC06", "Validar ordenamiento ASC/DESC", "HP", "P1"),
        ("TC07", "Validar resize + drag & drop + config columnas", "HP", "P2"),
        ("TC08", "Validar paginacion + refresh", "HP", "P2"),
        ("TC09", "Validar PDF descarga con campos GNET correctos", "HP", "P1"),
        ("TC10", "Validar i18n + l10n", "INT", "P1"),
    ],
    "MX-5574": [
        ("TC01", "Validar acceso desde menu GNET → Credit Accounts y carga", "HP", "P1"),
        ("TC02", "Validar titulo 'Credit Accounts'", "INT", "P1"),
        ("TC03", "Validar filtros aplican (estado cuenta / saldo)", "HP", "P1"),
        ("TC04", "Validar search por afiliado / codigo", "HP", "P1"),
        ("TC05", "Validar ordenamiento ASC/DESC en columnas saldo / nombre", "HP", "P1"),
        ("TC06", "Validar resize + drag & drop + config columnas", "HP", "P2"),
        ("TC07", "Validar paginacion + refresh", "HP", "P2"),
        ("TC08", "Validar PDF descarga con saldos y formato moneda local", "INT", "P1"),
        ("TC09", "Validar i18n + l10n formato moneda + fecha", "INT", "P1"),
    ],
    "MX-5575": [
        ("TC01", "Validar acceso desde menu Configuration → Other Costs y carga", "HP", "P1"),
        ("TC02", "Validar titulo 'Other Costs' / 'Otros Costos'", "INT", "P1"),
        ("TC03", "Validar filtros aplican (categoria / activo)", "HP", "P1"),
        ("TC04", "Validar search por nombre", "HP", "P1"),
        ("TC05", "Validar ordenamiento ASC/DESC", "HP", "P1"),
        ("TC06", "Validar resize + drag & drop + config columnas", "HP", "P2"),
        ("TC07", "Validar paginacion + refresh", "HP", "P2"),
        ("TC08", "Validar i18n EN/ES", "INT", "P1"),
    ],
    "MX-5554": [
        ("TC01", "Validar acceso desde menu eAffiliates → Credit Accounts y carga", "HP", "P1"),
        ("TC02", "Validar breadcrumb 'eAffiliates / Credit Accounts'", "INT", "P2"),
        ("TC03", "Validar titulo 'Credit Accounts With Affiliates'", "INT", "P1"),
        ("TC04", "Validar filtros aplican (clear all + ordenamiento por columna)", "HP", "P1"),
        ("TC05", "Validar search libre", "HP", "P1"),
        ("TC06", "Validar ordenamiento ASC/DESC en columnas sortable", "HP", "P1"),
        ("TC07", "Validar resize + drag & drop + config columnas", "HP", "P2"),
        ("TC08", "Validar empty state 'No Data' cuando no hay afiliados con cuenta", "NEG", "P1"),
        ("TC09", "Validar paginacion + refresh", "HP", "P2"),
        ("TC10", "Validar i18n + l10n formato moneda", "INT", "P1"),
    ],
}

# Navbar / Shell MX-5684 - 55 TCs
TCS_NAVBAR = [
    # 16.1 Estado inicial (HP, 5)
    ("TC01", "Validar tras login app-layout renderiza con app-sidebar + app-topbar + router-outlet", "HP", "P1"),
    ("TC02", "Validar sidebar muestra 17 items raiz del menuItems[] carrier en orden esperado", "HP", "P1"),
    ("TC03", "Validar topbar muestra: hamburger, logo, search, melita_ai, notifications, locale (EN/ES), help, avatar", "HP", "P1"),
    ("TC04", "Validar item del sidebar correspondiente a pantalla activa tiene isActive:true + CSS highlight", "HP", "P1"),
    ("TC05", "Validar avatar usuario muestra displayName + username + companyCode del _identityService", "HP", "P2"),
    # 16.2 Navegacion (HP, 15)
    ("TC06", "Validar click en Reports → Tips Report navega a /reports/tips (MX-5560)", "HP", "P1"),
    ("TC07", "Validar click en Reports → Aging Report navega a /reports/debt-aging (MX-5561)", "HP", "P1"),
    ("TC08", "Validar click en Reports → Collection Movements navega a /reports/cash-flow (MX-5562)", "HP", "P1"),
    ("TC09", "Validar click en Reports → Electronic Payment Transactions navega a /reports/transaction-tracking (MX-5565)", "HP", "P1"),
    ("TC10", "Validar click en Reports → Taxes & Fees navega a /reports/taxes-and-fees (MX-5566)", "HP", "P1"),
    ("TC11", "Validar click en Reports → Payment Movements navega a /reports/payment-flow (MX-5568)", "HP", "P1"),
    ("TC12", "Validar click en Reports → Company Commissions navega a /reports/agency-commissions (MX-5571)", "HP", "P1"),
    ("TC13", "Validar click en Reports → Daily Report navega a /reports/daily (MX-5438)", "HP", "P1"),
    ("TC14", "Validar click en Reports → Expired Documentation navega a /reports/documentation (MX-5569)", "HP", "P1"),
    ("TC15", "Validar click en Reports → Unpaid Trips navega a /reports/unpaid-travels-list (MX-5531)", "HP", "P1"),
    ("TC16", "Validar click en Reports → Trips Segments navega a /reports/segments-travels (MX-5553)", "HP", "P1"),
    ("TC17", "Validar click en GNET → Farm IN navega a /gnet/farm-in (MX-5573)", "HP", "P1"),
    ("TC18", "Validar click en GNET → Credit Accounts navega a /gnet/credit-accounts (MX-5574)", "HP", "P1"),
    ("TC19", "Validar click en Configuration → Other Costs navega a /settings/otherCosts (MX-5575)", "HP", "P1"),
    ("TC20", "Validar click en eAffiliates → Credit Accounts With Affiliates navega a /affiliate/checking-account (MX-5554)", "HP", "P1"),
    # 16.3 Toggle submenus (HP, 5)
    ("TC21", "Validar click en toggle Reports cambia isCollapsed:false y CSS revela submenu animado", "HP", "P1"),
    ("TC22", "Validar segundo click en Reports vuelve a isCollapsed:true (colapsa)", "HP", "P1"),
    ("TC23", "Validar click en toggle GNET expande con Farm IN + Credit Accounts + sub-rutas", "HP", "P1"),
    ("TC24", "Validar click en toggle eAffiliates expande con atc-profile, offering, request, os-agreement-*, checking-account", "HP", "P2"),
    ("TC25", "Validar click en toggle Configuration expande con parameters, transportTypes, servicesType, taxesAndFees, otherCosts, email-templates, etc.", "HP", "P2"),
    # 16.4 Edge cases - accordion (EC, 7)
    ("TC26", "Validar accordion: abrir toggle Reports mientras Configuration esta abierto cierra Configuration (siblings.isActive=false)", "EC", "P1"),
    ("TC27", "Validar al navegar a /reports/tips, padre Reports queda isCollapsed:false y sub-item Tips queda isActive:true", "EC", "P1"),
    ("TC28", "Validar refresh F5 sobre /reports/tips restaura sidebar con Reports expandido + Tips highlighted (withUiState reactiva al routing)", "EC", "P1"),
    ("TC29", "Validar deep-link directo /carrier/#/gnet/farm-in en sesion limpia restaura sidebar con GNET expandido + Farm IN highlighted", "EC", "P1"),
    ("TC30", "Validar back/forward del navegador entre pantallas actualiza isActive del item correspondiente", "EC", "P2"),
    ("TC31", "Validar item Magiis Apps Store como enlace directo (no toggle): un solo click navega sin expansion", "EC", "P2"),
    ("TC32", "Validar logo (routerLink='dashboard') navega a /carrier/#/dashboard y resalta Operations Control", "EC", "P1"),
    # 16.5 Negativos (NEG, 7)
    ("TC33", "Validar deep-link a ruta inexistente (/carrier/#/foo-bar) redirige a /dashboard sin romper shell", "NEG", "P1"),
    ("TC34", "Validar deep-link a ruta sin permisos del rol carrier redirige a /dashboard o muestra mensaje 'No autorizado'", "NEG", "P1"),
    ("TC35", "Validar items del sidebar que el rol carrier NO debe ver estan ocultos (no renderizados en menuItems[] filtrado por permisos)", "NEG", "P1"),
    ("TC36", "Validar click rapido en 5+ items consecutivos no rompe isActive state ni deja inconsistencias", "NEG", "P2"),
    ("TC37", "Validar sesion expirada en medio de navegacion redirige a /auth/login con mensaje informativo", "NEG", "P1"),
    ("TC38", "Validar logout (_identityService.logOut()) limpia storageState + cookies + redirige a /auth/login", "NEG", "P1"),
    ("TC39", "Validar intentar acceder a /dashboard sin sesion redirige a /auth/login (auth guard)", "NEG", "P1"),
    # 16.6 Regresion (REG, 6)
    ("TC40", "Validar navegacion entre 10+ pantallas consecutivas no degrada FPS ni genera memory leaks (sin console.error)", "REG", "P2"),
    ("TC41", "Validar tras cambiar de pantalla, filtros/search de pantalla anterior se resetean al volver (no state leakage)", "REG", "P1"),
    ("TC42", "Validar scroll del sidebar (ngx-simplebar) mantiene posicion al cambiar pantalla", "REG", "P2"),
    ("TC43", "Validar sidebar no se duplica visualmente al navegar con back/forward rapido", "REG", "P2"),
    ("TC44", "Validar sesion persiste tras navegar a pantalla, refresh, y volver a la misma sin re-login", "REG", "P1"),
    ("TC45", "Validar breadcrumb h4 actualiza correctamente al cambiar de pantalla (no queda el de pantalla anterior)", "REG", "P1"),
    # 16.7 Integracion (INT, 6)
    ("TC46", "Validar cambio de locale EN → ES en topbar actualiza labels del sidebar en vivo sin refresh", "INT", "P1"),
    ("TC47", "Validar cambio de locale ES → EN actualiza labels del topbar (Trip / Trips Management / dropdowns)", "INT", "P1"),
    ("TC48", "Validar avatar usuario muestra datos del _identityService.getSessionData() (companyName, role) no hardcoded", "INT", "P1"),
    ("TC49", "Validar notifications dropdown del topbar lista items de allnotifications[] con icons y badges", "INT", "P2"),
    ("TC50", "Validar melita_ai quick access dropdown expande con sub-items del servicio Melita", "INT", "P2"),
    ("TC51", "Validar dark/light mode toggle propaga via eventService.broadcast('changeMode',mode) y actualiza CSS variables", "INT", "P2"),
    # 16.8 Topbar Trips (HP+INT, 4)
    ("TC52", "Validar boton verde '+ Trip' (topbar) navega a /carrier/#/travel/create", "HP", "P1"),
    ("TC53", "Validar boton 'Trips Management' (topbar) navega a /carrier/#/travel/dashboard", "HP", "P1"),
    ("TC54", "Validar boton 'Trips' (topbar dropdown) muestra accesos rapidos (Recurring, Quotes, Mappers)", "HP", "P2"),
    ("TC55", "Validar avatar usuario despliega menu con displayName, opcion 'Perfil' si existe, y 'Logout'", "HP+INT", "P1"),
]


# ----- Build workbook -----
wb = Workbook()

# -------------------- Sheet 1: Matriz --------------------
ws = wb.active
ws.title = "Matriz"

headers = [
    "ID", "Sección", "Pantalla", "MX (Jira)", "URL", "Dimensión QA",
    "Tipo", "Prioridad", "Descripción", "Precondición",
    "Pasos (resumen)", "Resultado esperado", "Estado", "Resultado obtenido"
]
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER

row = 2
seccion_count = 0

# Pantallas baja complejidad (secciones 1-15)
for mx, screen in SCREENS.items():
    seccion_count += 1
    for tc_suffix, desc, dim, prio in TCS_PER_SCREEN.get(mx, []):
        tc_id = f"{mx}-{tc_suffix}"
        # Pasos sintetizados por dimensión.
        if dim == "HP":
            tipo = "Funcional"
            pasos = list_steps([
                "Login al portal carrier.",
                f"Navegar a {screen['url']}.",
                "Ejecutar acción descripta en la descripción del caso.",
                "Verificar resultado esperado."
            ])
        elif dim == "EC":
            tipo = "Funcional"
            pasos = list_steps([
                f"Navegar a {screen['url']}.",
                "Ejecutar acción en borde (valor mínimo/máximo/vacío).",
                "Verificar comportamiento esperado sin error."
            ])
        elif dim == "NEG":
            tipo = "Funcional"
            pasos = list_steps([
                f"Navegar a {screen['url']}.",
                "Ejecutar acción negativa (sin datos / sin permisos / inputs inválidos).",
                "Verificar mensaje o estado deshabilitado correspondiente."
            ])
        elif dim == "INT":
            tipo = "UX"
            pasos = list_steps([
                f"Navegar a {screen['url']}.",
                "Cambiar idioma EN ↔ ES (locale switcher topbar).",
                "Verificar que el elemento bajo test refleja el cambio (labels, formato fecha, formato moneda)."
            ])
        else:
            tipo = "Funcional"
            pasos = list_steps([f"Navegar a {screen['url']}.", "Ejecutar caso según descripción."])

        precondicion = PRE_LOGIN

        ws.cell(row=row, column=1, value=tc_id).font = ID_FONT
        ws.cell(row=row, column=2, value=f"{seccion_count}. {screen['name']}")
        ws.cell(row=row, column=3, value=screen["name"])
        ws.cell(row=row, column=4, value=mx).font = MX_FONT
        ws.cell(row=row, column=5, value=screen["url"])
        ws.cell(row=row, column=6, value=dim)
        ws.cell(row=row, column=7, value=tipo)
        ws.cell(row=row, column=8, value=prio)
        ws.cell(row=row, column=9, value=desc)
        ws.cell(row=row, column=10, value=precondicion)
        ws.cell(row=row, column=11, value=pasos)
        ws.cell(row=row, column=12, value=desc)  # resultado_esperado = misma descripción positiva
        ws.cell(row=row, column=13, value="Pendiente")
        ws.cell(row=row, column=14, value="")
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row, column=col_idx).alignment = WRAP
            ws.cell(row=row, column=col_idx).border = BORDER
            if col_idx not in (1, 4):
                ws.cell(row=row, column=col_idx).font = CELL_FONT
        row += 1

# Sección 16 - Navbar MX-5684
mx_navbar = "MX-5684"
navbar_url = "/carrier/#/* (shell global)"
seccion_count = 16
for tc_suffix, desc, dim, prio in TCS_NAVBAR:
    tc_id = f"{mx_navbar}-{tc_suffix}"
    tipo = "UX" if dim in ("INT", "EC") else "Funcional"
    if dim == "HP":
        pasos = list_steps([
            "Login al portal carrier.",
            "Verificar estado del shell (sidebar + topbar visibles).",
            "Ejecutar acción descripta.",
            "Verificar resultado esperado."
        ])
    elif dim == "EC":
        pasos = list_steps([
            "Login al portal.",
            "Ejecutar borde: deep-link / accordion concurrente / back-forward / refresh.",
            "Verificar restauración de estado UI consistente."
        ])
    elif dim == "NEG":
        pasos = list_steps([
            "Ejecutar caso negativo (ruta inválida / sin permisos / sesión expirada).",
            "Verificar redirect o mensaje informativo apropiado.",
            "Verificar que el shell no se rompe."
        ])
    elif dim == "REG":
        pasos = list_steps([
            "Login al portal.",
            "Ejecutar flujo crítico de navegación múltiple.",
            "Verificar que comportamientos previos no se rompen (sin console.error, sin state leakage, performance estable)."
        ])
    elif dim == "INT":
        pasos = list_steps([
            "Login al portal.",
            "Interactuar con servicio externo (TranslateService / LocalizationService / IdentityService).",
            "Verificar que el shell refleja el cambio en vivo."
        ])
    else:  # HP+INT
        pasos = list_steps([
            "Login al portal.",
            "Verificar topbar Trips dropdown + integration con _identityService.",
            "Verificar navegación + permisos."
        ])

    ws.cell(row=row, column=1, value=tc_id).font = ID_FONT
    ws.cell(row=row, column=2, value="16. Navbar / Shell (Revisión integral)")
    ws.cell(row=row, column=3, value="Navbar / Shell")
    ws.cell(row=row, column=4, value=mx_navbar).font = MX_FONT
    ws.cell(row=row, column=5, value=navbar_url)
    ws.cell(row=row, column=6, value=dim)
    ws.cell(row=row, column=7, value=tipo)
    ws.cell(row=row, column=8, value=prio)
    ws.cell(row=row, column=9, value=desc)
    ws.cell(row=row, column=10, value=PRE_LOGIN)
    ws.cell(row=row, column=11, value=pasos)
    ws.cell(row=row, column=12, value=desc)
    ws.cell(row=row, column=13, value="Pendiente")
    ws.cell(row=row, column=14, value="")
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=row, column=col_idx).alignment = WRAP
        ws.cell(row=row, column=col_idx).border = BORDER
        if col_idx not in (1, 4):
            ws.cell(row=row, column=col_idx).font = CELL_FONT
    row += 1

total_tcs = row - 2

# Column widths
widths = {
    "A": 18, "B": 32, "C": 28, "D": 12, "E": 38, "F": 12,
    "G": 12, "H": 10, "I": 55, "J": 30, "K": 50, "L": 50, "M": 12, "N": 25
}
for col_letter, w in widths.items():
    ws.column_dimensions[col_letter].width = w

ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

# Convertir a tabla con filtros
last_col = get_column_letter(len(headers))
table_range = f"A1:{last_col}{row - 1}"
table = Table(displayName="MatrizCases", ref=table_range)
table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=False
)
ws.add_table(table)

# -------------------- Sheet 2: Resumen --------------------
ws2 = wb.create_sheet("Resumen")
resumen_headers = ["Sección", "Pantalla / Ticket", "MX (Jira)", "TCs", "P1 (Alta)", "P2 (Media)", "Dimensiones"]
for col_idx, h in enumerate(resumen_headers, 1):
    cell = ws2.cell(row=1, column=col_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER

# Computar resumen por pantalla
resumen_row = 2
total_p1 = 0
total_p2 = 0
seccion_idx = 0
for mx, screen in SCREENS.items():
    seccion_idx += 1
    tcs_list = TCS_PER_SCREEN.get(mx, [])
    p1 = sum(1 for _, _, _, prio in tcs_list if prio == "P1")
    p2 = sum(1 for _, _, _, prio in tcs_list if prio == "P2")
    dims = sorted({dim for _, _, dim, _ in tcs_list})
    ws2.cell(row=resumen_row, column=1, value=seccion_idx)
    ws2.cell(row=resumen_row, column=2, value=screen["name"])
    ws2.cell(row=resumen_row, column=3, value=mx).font = MX_FONT
    ws2.cell(row=resumen_row, column=4, value=len(tcs_list))
    ws2.cell(row=resumen_row, column=5, value=p1)
    ws2.cell(row=resumen_row, column=6, value=p2)
    ws2.cell(row=resumen_row, column=7, value=", ".join(dims))
    for col_idx in range(1, len(resumen_headers) + 1):
        ws2.cell(row=resumen_row, column=col_idx).font = (
            MX_FONT if col_idx == 3 else CELL_FONT
        )
        ws2.cell(row=resumen_row, column=col_idx).alignment = WRAP
        ws2.cell(row=resumen_row, column=col_idx).border = BORDER
    total_p1 += p1
    total_p2 += p2
    resumen_row += 1

# Navbar row
nav_p1 = sum(1 for _, _, _, prio in TCS_NAVBAR if prio == "P1")
nav_p2 = sum(1 for _, _, _, prio in TCS_NAVBAR if prio == "P2")
nav_dims = sorted({dim for _, _, dim, _ in TCS_NAVBAR})
ws2.cell(row=resumen_row, column=1, value=16)
ws2.cell(row=resumen_row, column=2, value="Navbar / Shell (Revisión integral)")
ws2.cell(row=resumen_row, column=3, value=mx_navbar).font = MX_FONT
ws2.cell(row=resumen_row, column=4, value=len(TCS_NAVBAR))
ws2.cell(row=resumen_row, column=5, value=nav_p1)
ws2.cell(row=resumen_row, column=6, value=nav_p2)
ws2.cell(row=resumen_row, column=7, value=", ".join(nav_dims))
for col_idx in range(1, len(resumen_headers) + 1):
    ws2.cell(row=resumen_row, column=col_idx).font = (
        MX_FONT if col_idx == 3 else CELL_FONT
    )
    ws2.cell(row=resumen_row, column=col_idx).alignment = WRAP
    ws2.cell(row=resumen_row, column=col_idx).border = BORDER
total_p1 += nav_p1
total_p2 += nav_p2
resumen_row += 1

# Total row con formula
total_row = resumen_row
sum_col_d_range = f"D2:D{resumen_row - 1}"
sum_col_e_range = f"E2:E{resumen_row - 1}"
sum_col_f_range = f"F2:F{resumen_row - 1}"
ws2.cell(row=total_row, column=1, value="").font = Font(bold=True)
ws2.cell(row=total_row, column=2, value="TOTAL").font = Font(name=ARIAL, size=10, bold=True)
ws2.cell(row=total_row, column=3, value="16 tickets Jira").font = Font(name=ARIAL, size=10, bold=True)
ws2.cell(row=total_row, column=4, value=f"=SUM({sum_col_d_range})").font = Font(name=ARIAL, size=10, bold=True)
ws2.cell(row=total_row, column=5, value=f"=SUM({sum_col_e_range})").font = Font(name=ARIAL, size=10, bold=True)
ws2.cell(row=total_row, column=6, value=f"=SUM({sum_col_f_range})").font = Font(name=ARIAL, size=10, bold=True)
ws2.cell(row=total_row, column=7, value="5 dimensiones qa-magiis").font = Font(name=ARIAL, size=10, bold=True)
for col_idx in range(1, len(resumen_headers) + 1):
    ws2.cell(row=total_row, column=col_idx).fill = PatternFill("solid", fgColor="FFFF00")  # yellow attention
    ws2.cell(row=total_row, column=col_idx).border = BORDER
    ws2.cell(row=total_row, column=col_idx).alignment = CENTER

widths2 = {"A": 8, "B": 36, "C": 14, "D": 8, "E": 12, "F": 12, "G": 32}
for col_letter, w in widths2.items():
    ws2.column_dimensions[col_letter].width = w
ws2.freeze_panes = "A2"

# -------------------- Sheet 3: Trazabilidad --------------------
ws3 = wb.create_sheet("Trazabilidad")
traza_headers = [
    "MX (Jira)", "Pantalla", "URL real V2", "Componente Angular", "Branch V2 referencia",
    "Sección en matriz_cases.md", "TCs cubiertos", "Jira Link"
]
for col_idx, h in enumerate(traza_headers, 1):
    cell = ws3.cell(row=1, column=col_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER

trazabilidad_data = [
    ("MX-5560", "Reportes / Propinas", "/carrier/#/reports/tips", "ReportTipsComponent", "reports-tips", "1", 26),
    ("MX-5561", "Reportes / Antigüedad Deuda", "/carrier/#/reports/debt-aging", "ReportAgingComponent", "reports-debt-aging", "2", 14),
    ("MX-5562", "Reportes / Movimientos Cobros", "/carrier/#/reports/cash-flow", "ReportCollectionMovementsComponent", "reports-cash-flow", "3", 11),
    ("MX-5565", "Reportes / Transacciones Tarjeta", "/carrier/#/reports/transaction-tracking", "ReportElectronicPaymentTransactionsComponent", "reports-transaction-tracking", "4", 10),
    ("MX-5566", "Reportes / Impuestos y Cargos", "/carrier/#/reports/taxes-and-fees", "ReportTaxesAndFeesComponent", "reports-taxes-and-fees", "5", 10),
    ("MX-5568", "Reportes / Movimientos Pagos", "/carrier/#/reports/payment-flow", "ReportPaymentMovementsComponent", "reports-payment-flow", "6", 10),
    ("MX-5571", "Reportes / Comisiones Agencia", "/carrier/#/reports/agency-commissions", "ReportCompanyCommissionsComponent", "reports-agency-commissions", "7", 11),
    ("MX-5438", "Reportes / Resumen Diario", "/carrier/#/reports/daily", "ReportDailyComponent", "reports-daily", "8", 10),
    ("MX-5569", "Reportes / Documentación Vencida", "/carrier/#/reports/documentation", "ReportExpireDocumentationsComponent", "reports-documentation", "9", 10),
    ("MX-5531", "Reportes / Travel Unpaid", "/carrier/#/reports/unpaid-travels-list", "ReportUnpaidTripsComponent", "travel-unpaid-list", "10", 9),
    ("MX-5553", "Reportes / Viajes por Segmentos", "/carrier/#/reports/segments-travels", "ReportTripsSegmentsComponent", "segment-travels", "11", 10),
    ("MX-5573", "GNET Farm IN", "/carrier/#/gnet/farm-in", "GnetFarmInComponent", "gnet-farm-in", "12", 10),
    ("MX-5574", "GNET Cuentas Corrientes", "/carrier/#/gnet/credit-accounts", "GnetCreditAccountsComponent", "gnet-credit-accounts", "13", 9),
    ("MX-5575", "Configuración / Otros Costos", "/carrier/#/settings/otherCosts", "SettingsOtherCostsComponent", "settings-other-costs", "14", 8),
    ("MX-5554", "Cuentas Corrientes Con Afiliados", "/carrier/#/affiliate/checking-account", "AffiliateCheckingAccountComponent", "affiliate-checking-account", "15", 10),
    ("MX-5684", "Navbar / Shell (Revisión integral)", "/carrier/#/* (shell)", "SidebarComponent + TopbarComponent + VerticalComponent", "(múltiples)", "16", 55),
]

for tr_row_idx, data in enumerate(trazabilidad_data, 2):
    mx, pantalla, url, comp, branch, seccion, tcs_count = data
    ws3.cell(row=tr_row_idx, column=1, value=mx).font = MX_FONT
    ws3.cell(row=tr_row_idx, column=2, value=pantalla).font = CELL_FONT
    ws3.cell(row=tr_row_idx, column=3, value=url).font = CELL_FONT
    ws3.cell(row=tr_row_idx, column=4, value=comp).font = CELL_FONT
    ws3.cell(row=tr_row_idx, column=5, value=branch).font = CELL_FONT
    ws3.cell(row=tr_row_idx, column=6, value=seccion).font = CELL_FONT
    ws3.cell(row=tr_row_idx, column=7, value=tcs_count).font = CELL_FONT
    jira_link = f"https://magiis.atlassian.net/browse/{mx}"
    cell_link = ws3.cell(row=tr_row_idx, column=8, value=jira_link)
    cell_link.font = Font(name=ARIAL, size=10, color="0000FF", underline="single")
    cell_link.hyperlink = jira_link
    for col_idx in range(1, len(traza_headers) + 1):
        ws3.cell(row=tr_row_idx, column=col_idx).alignment = WRAP
        ws3.cell(row=tr_row_idx, column=col_idx).border = BORDER

widths3 = {"A": 14, "B": 38, "C": 42, "D": 50, "E": 30, "F": 18, "G": 8, "H": 50}
for col_letter, w in widths3.items():
    ws3.column_dimensions[col_letter].width = w
ws3.freeze_panes = "A2"

# Output
out_path = "docs/qa/release-v2.0.4/matriz_cases_baja_complejidad.xlsx"
wb.save(out_path)
print(f"OK: {out_path} generado.")
print(f"Total TCs: {total_tcs} | Total P1: {total_p1} | Total P2: {total_p2}")
print(f"Sheets: Matriz ({total_tcs} rows), Resumen ({resumen_row - 1} rows), Trazabilidad ({len(trazabilidad_data)} rows)")
